from __future__ import annotations

import io
import random
from dataclasses import dataclass
from hashlib import sha256
from typing import Any
from uuid import UUID

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.core.config import get_settings
from app.models.report import ReportFormat, ReportType


settings = get_settings()


@dataclass
class GeneratedReport:
    file_name: str
    content_type: str
    payload: bytes
    checksum: str
    row_count: int


class ReportBuilderService:
    def build_report(
        self,
        *,
        report_id: UUID,
        report_type: ReportType,
        report_format: ReportFormat,
        filters: dict[str, Any],
    ) -> GeneratedReport:
        rows = self._build_rows(report_id=report_id, report_type=report_type, filters=filters)
        if report_format == ReportFormat.excel:
            payload = self._build_excel(report_type=report_type, rows=rows, filters=filters)
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            extension = "xlsx"
        else:
            payload = self._build_pdf(report_type=report_type, rows=rows, filters=filters)
            content_type = "application/pdf"
            extension = "pdf"

        file_name = f"{report_type.value}_{str(report_id)[:8]}.{extension}"
        checksum = sha256(payload).hexdigest()
        return GeneratedReport(
            file_name=file_name,
            content_type=content_type,
            payload=payload,
            checksum=checksum,
            row_count=len(rows),
        )

    def _build_rows(self, *, report_id: UUID, report_type: ReportType, filters: dict[str, Any]) -> list[dict[str, Any]]:
        seed = int(str(report_id.int)[-8:])
        rng = random.Random(seed)
        area = filters.get("area") or "General"
        category = filters.get("category") or "Core"
        base_rows = {
            ReportType.sales_summary: 18,
            ReportType.operations_kpis: 14,
            ReportType.audit_log: 25,
        }[report_type]
        rows: list[dict[str, Any]] = []
        for index in range(1, base_rows + 1):
            rows.append(
                {
                    "period": f"Lote {index:02d}",
                    "area": area,
                    "category": category,
                    "owner": filters.get("requested_user") or f"user_{index:02d}",
                    "status": filters.get("status") or rng.choice(["open", "closed", "review"]),
                    "amount": round(rng.uniform(1500, 12000), 2),
                    "volume": rng.randint(10, 120),
                    "score": round(rng.uniform(75, 99), 1),
                }
            )
        return rows

    def _build_excel(self, *, report_type: ReportType, rows: list[dict[str, Any]], filters: dict[str, Any]) -> bytes:
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Resumen"
        data_sheet = workbook.create_sheet("Datos")

        summary_sheet["A1"] = "Reporte corporativo"
        summary_sheet["A1"].font = Font(size=16, bold=True)
        summary_sheet["A3"] = "Tipo"
        summary_sheet["B3"] = report_type.value
        summary_sheet["A4"] = "Rango"
        summary_sheet["B4"] = f"{filters['start_date']} a {filters['end_date']}"
        summary_sheet["A5"] = "Filas"
        summary_sheet["B5"] = len(rows)
        summary_sheet["A6"] = "Moneda"
        summary_sheet["B6"] = settings.default_currency

        dataframe = pd.DataFrame(rows)
        headers = list(dataframe.columns)
        data_sheet.append(headers)
        for item in dataframe.to_dict(orient="records"):
            data_sheet.append(list(item.values()))

        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        for column_index, header in enumerate(headers, 1):
            cell = data_sheet.cell(row=1, column=column_index)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            width = max(len(header), 12)
            data_sheet.column_dimensions[get_column_letter(column_index)].width = width + 6

        stream = io.BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    def _build_pdf(self, *, report_type: ReportType, rows: list[dict[str, Any]], filters: dict[str, Any]) -> bytes:
        stream = io.BytesIO()
        document = SimpleDocTemplate(
            stream,
            pagesize=landscape(A4),
            leftMargin=24,
            rightMargin=24,
            topMargin=28,
            bottomMargin=24,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("TitleAccent", parent=styles["Title"], textColor=colors.HexColor("#16324F"))
        story = [
            Paragraph("Reporte corporativo", title_style),
            Spacer(1, 8),
            Paragraph(f"Tipo: {report_type.value}", styles["Heading3"]),
            Paragraph(f"Periodo: {filters['start_date']} a {filters['end_date']}", styles["BodyText"]),
            Spacer(1, 12),
        ]

        headers = ["period", "area", "category", "owner", "status", "amount", "volume", "score"]
        table_data = [headers]
        for row in rows:
            table_data.append(
                [
                    row["period"],
                    row["area"],
                    row["category"],
                    row["owner"],
                    row["status"],
                    f"{row['amount']:.2f}",
                    str(row["volume"]),
                    f"{row['score']:.1f}",
                ]
            )

        table = Table(table_data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16324F")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B0BEC5")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
                    ("ALIGN", (5, 1), (-1, -1), "RIGHT"),
                ]
            )
        )
        story.append(table)

        document.build(story)
        return stream.getvalue()
